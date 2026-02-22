import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Optional
from uuid import UUID

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from structlog import get_logger

from app.domain.entities import FinanceServiceType, PaymentMethod, Transaction

from .interfaces import ExcelLoader

logger = get_logger(__name__)


class TransactionExcelLoader(ExcelLoader[Transaction]):
    """
    Loads transactions from Excel file safely using openpyxl.

    Expected columns (first row must be headers):
        transaction_id, client_id, transaction_date, service,
        amount, payment_method, city, consultant
    """

    REQUIRED_COLUMNS = {
        "transaction_id",
        "client_id",
        "transaction_date",
        "service",
        "amount",
        "payment_method",
        "city",
        "consultant",
    }

    MAX_FILE_SIZE = 100 * 1024 * 1024
    MAX_ROWS = 1_000_000
    MAX_COLUMNS = 50
    ALLOWED_EXTENSIONS = {".xlsx"}

    def __init__(self, sheet_name: str | int = 0, header_row: int = 1):
        """
        Initialize Excel loader.

        Args:
            sheet_name: Excel sheet name or index to read.
            header_row: Row number containing column headers (1-based).
        """
        self.sheet_name = sheet_name
        self.header_row = header_row
        self._validate_sheet_name()

    def _validate_sheet_name(self) -> None:
        if isinstance(self.sheet_name, str):
            if len(self.sheet_name) > 31:
                raise ValueError(f"Sheet name too long: {self.sheet_name}")

            if not re.match(r"^[a-zA-Z0-9\s_-]{1,31}$", self.sheet_name):
                raise ValueError(f"Invalid sheet name: {self.sheet_name}")

            if (
                ".." in self.sheet_name
                or "/" in self.sheet_name
                or "\\" in self.sheet_name
            ):
                raise ValueError(
                    "Invalid sheet name (path traversal attempt): "
                    f"{self.sheet_name}"
                )

    def _validate_file(self, source: Path) -> None:
        """
        Validate file path for security.

        Raises:
            FileNotFoundError: If file doesn't exist.
            ValueError: If file is too large or has invalid extension.
        """
        if not source.exists():
            raise FileNotFoundError(f"Excel file not found: {source}")

        if source.suffix.lower() not in self.ALLOWED_EXTENSIONS:
            raise ValueError(
                f"Invalid file extension: {source.suffix}. "
                f"Allowed: {self.ALLOWED_EXTENSIONS}"
            )

        file_size = source.stat().st_size
        if file_size > self.MAX_FILE_SIZE:
            raise ValueError(
                f"File too large: {file_size} bytes "
                f"(max: {self.MAX_FILE_SIZE} bytes)"
            )

    def supports(self, source: Path) -> bool:
        """Check if file is Excel with additional validation."""
        try:
            if source.suffix.lower() not in self.ALLOWED_EXTENSIONS:
                return False

            if source.exists() and source.stat().st_size > 0:
                with open(source, "rb") as f:
                    header = f.read(4)
                    excel_signatures = {
                        b"\x50\x4b\x03\x04",  # .xlsx
                    }
                    if header not in excel_signatures:
                        return False

                    with zipfile.ZipFile(source) as zf:
                        for name in zf.namelist():
                            if any(
                                ext in name.lower()
                                for ext in [".exe", ".dll", ".vbs", ".ps1"]
                            ):
                                return False
            return True
        except Exception:
            return False

    def _get_worksheet(self, workbook) -> Worksheet:
        """Get worksheet by name or index safely."""
        if isinstance(self.sheet_name, str):
            if self.sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Sheet '{self.sheet_name}' not found. "
                    f"Available: {workbook.sheetnames}"
                )
            return workbook[self.sheet_name]
        else:
            if self.sheet_name >= len(workbook.worksheets):
                raise ValueError(
                    f"Sheet index {self.sheet_name} out of range. "
                    f"Max index: {len(workbook.worksheets) - 1}"
                )
            return workbook.worksheets[self.sheet_name]

    def _get_headers(self, worksheet: Worksheet) -> list[str]:
        """Extract headers from the specified row."""
        headers = []
        for col_idx in range(
            1, min(worksheet.max_column + 1, self.MAX_COLUMNS + 1)
        ):
            cell = worksheet.cell(row=self.header_row, column=col_idx)
            cell_value = self._safe_get_cell_value(cell)

            if cell_value is None:
                break
            headers.append(str(cell_value).strip().lower())
        return headers

    def load(self, source: Path) -> Iterator[Transaction]:
        """
        Load transactions from Excel file safely using openpyxl.

        Args:
            source: Path to Excel file.

        Yields:
            Transaction entities.

        Raises:
            FileNotFoundError: If file doesn't exist.
            ValueError: If required columns are missing or file invalid.
        """
        self._validate_file(source)

        logger.info(f"Loading transactions from {source.name}")

        workbook = None
        try:
            workbook = load_workbook(
                filename=source,
                read_only=True,
                data_only=True,
                keep_links=False,
                keep_vba=False,
                rich_text=False,
            )

            worksheet = self._get_worksheet(workbook)

            if worksheet.max_row < self.header_row:
                logger.warning("Worksheet is empty or has no header row")
                return

            if worksheet.max_row - self.header_row > self.MAX_ROWS:
                raise ValueError(
                    f"Too many rows: {worksheet.max_row - self.header_row} "
                    f"(max: {self.MAX_ROWS})"
                )

            headers = self._get_headers(worksheet)
            if not headers:
                raise ValueError("No headers found in the first row")

            missing_cols = self.REQUIRED_COLUMNS - set(headers)
            if missing_cols:
                raise ValueError(
                    f"Missing required columns: {missing_cols}. "
                    f"Found headers: {headers}"
                )

            column_mapping = {}
            for col_idx, header in enumerate(headers, start=1):
                if header in self.REQUIRED_COLUMNS:
                    column_mapping[header] = col_idx

            logger.info(
                f"Loading rows from Excel. "
                f"Total rows: {worksheet.max_row - self.header_row}"
            )

            rows_processed = 0
            current_row = 1

            for row in worksheet.iter_rows(
                min_row=self.header_row + 1, values_only=True
            ):
                try:
                    if all(cell is None for cell in row):
                        continue

                    row_data = {}
                    for header, col_idx in column_mapping.items():
                        if col_idx - 1 < len(row):
                            row_data[header] = row[col_idx - 1]

                    transaction = self._row_to_transaction(row_data)
                    if transaction:
                        yield transaction
                        rows_processed += 1

                    current_row += 1

                except Exception as e:
                    logger.warning(f"Failed to parse row {current_row}: {e}")
                    continue

            logger.info(f"Successfully loaded {rows_processed} transactions")

        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise
        finally:
            if workbook:
                workbook.close()

    def load_sheet(
        self, source: Path, sheet_name: str | int
    ) -> Iterator[Transaction]:
        """Load from specific sheet with validation."""
        if isinstance(sheet_name, str):
            if len(sheet_name) > 31:
                raise ValueError(f"Sheet name too long: {sheet_name}")

        old_sheet = self.sheet_name
        self.sheet_name = sheet_name
        try:
            yield from self.load(source)
        finally:
            self.sheet_name = old_sheet

    def _row_to_transaction(
        self, row_data: dict[str, Any]
    ) -> Optional[Transaction]:
        """
        Convert Excel row data to Transaction entity.
        """
        try:
            for field in ["transaction_id", "client_id"]:
                value = row_data.get(field)
                if value and len(str(value)) > 100:
                    logger.debug(f"Field {field} too long, skipping")
                    return None

            transaction_id = self._parse_uuid(row_data.get("transaction_id"))
            client_id = self._parse_uuid(row_data.get("client_id"))

            transaction_date = self._parse_date(
                row_data.get("transaction_date")
            )
            amount = self._parse_amount(row_data.get("amount"))

            if amount is None:
                logger.debug(
                    f"Skipping transaction {transaction_id}: invalid amount."
                )
                return None

            raw_service = self._clean_string(row_data.get("service"))
            raw_payment_method = self._clean_string(
                row_data.get("payment_method")
            )
            city = self._clean_string(row_data.get("city"))
            consultant = self._clean_string(row_data.get("consultant"))

            return Transaction(
                id=transaction_id,
                client_id=client_id,
                transaction_date=transaction_date,
                raw_service=raw_service,
                amount=amount,
                raw_payment_method=raw_payment_method,
                city=city,
                consultant=consultant,
            )

        except Exception as e:
            logger.warning(f"Failed to parse row: {e}")
            return None

    def _parse_uuid(self, value: Any) -> Optional[UUID]:
        """Parse UUID from various formats safely."""
        if value is None or value == "":
            return None

        try:
            str_value = str(value).strip()

            if len(str_value) > 36:
                logger.debug(f"Value too long for UUID: {str_value[:50]}...")
                return None

            if not re.match(r"^[0-9a-f-]+$", str_value, re.I):
                logger.debug(f"Invalid characters in UUID: {str_value}")
                return None

            return UUID(str_value)
        except (ValueError, AttributeError):
            logger.debug(f"Invalid UUID: {value}")
            return None

    def _parse_date(self, value: Any) -> Optional[datetime]:
        """Parse date with fallback."""
        if value is None or value in ("INVALID_DATE", ""):
            return None

        try:
            if isinstance(value, datetime):
                if value.year < 1900 or value.year > 2100:
                    logger.debug(f"Date out of reasonable range: {value}")
                    return None
                return value

            if isinstance(value, str):
                for fmt in [
                    "%Y-%m-%d",
                    "%d.%m.%Y",
                    "%d/%m/%Y",
                    "%Y/%m/%d",
                    "%d-%m-%Y",
                ]:
                    try:
                        date = datetime.strptime(value.strip(), fmt)
                        if 1900 <= date.year <= 2100:
                            return date
                    except ValueError:
                        continue

            date = pd.to_datetime(value, errors="coerce")
            if pd.notna(date):
                if 1900 <= date.year <= 2100:
                    return date
            return None

        except Exception:
            return None

    def _parse_amount(self, value: Any) -> Optional[float]:
        """Parse amount with comma decimal separator safely."""
        if value is None:
            return None

        try:
            str_value = str(value).strip().replace(" ", "")

            if len(str_value) > 50:
                logger.debug(f"Amount string too long: {len(str_value)} chars")
                return None

            if "," in str_value and "." not in str_value:
                str_value = str_value.replace(",", ".")

            str_value = re.sub(r"[^\d.,-]", "", str_value)

            amount = float(str_value)

            if amount > 1_000_000_000_000:
                logger.debug(f"Amount suspiciously large: {amount}")
                return None

            return amount if amount > 0 else None

        except (ValueError, TypeError):
            return None

    def _clean_string(self, value: Any) -> str:
        """Clean string value safely."""
        if value is None:
            return "EMPTY"

        cleaned = str(value).strip()

        if len(cleaned) > 1000:
            logger.debug(
                f"String length suspiciously large: {len(cleaned)} chars"
            )

        return cleaned if cleaned else "EMPTY"

    def _safe_get_cell_value(self, cell) -> str:
        """Formulas escape"""
        if cell.data_type == "f":
            logger.warning(
                f"Formula detected in cell {cell.coordinate}, "
                "treating as empty"
            )
            return ""
        return str(cell.value) if cell.value is not None else ""

    def _to_service_enum(self, value: str) -> FinanceServiceType:
        """Convert string to FinanceServiceType enum."""
        try:
            return FinanceServiceType(value)
        except ValueError:
            logger.debug(f"Unknown service type: {value}, using UNKNOWN")
            return FinanceServiceType.UNKNOWN

    def _to_payment_enum(self, value: str) -> PaymentMethod:
        """Convert string to PaymentMethod enum."""
        try:
            return PaymentMethod(value)
        except ValueError:
            logger.debug(f"Unknown payment method: {value}, using UNKNOWN")
            return PaymentMethod.UNKNOWN
